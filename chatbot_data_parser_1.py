#! /usr/bin/env python

import argparse
import openpyxl


def compare_data(main_data, sub_data):
    wb = openpyxl.Workbook()
    ws1 = wb.create_sheet("사용자 의도", 0)
    ws2 = wb.create_sheet("사용자 예상질문", 1)
    ws3 = wb.create_sheet("매칭 안된 사용자 의도", 2)
    ws4 = wb.create_sheet("매칭 안된 사용자 예상질문", 3)

    ws1_count = 0
    ws2_count = 0
    ws3_count = 0
    ws4_count = 0

    for mk, mv in main_data.items():
        for sk, sv in sub_data.items():
            if mk in sv:
                sv.insert(0, sv.pop(sv.index(mk)))
                ws1_count += 1
                ws1['A' + str(ws1_count)] = mk
                ws1['B' + str(ws1_count)] = mv
                for i in sv:
                    ws2_count += 1
                    ws2['A' + str(ws2_count)] = i
                ws2_count += 1
                ws2['A' + str(ws2_count)] = ""
                sub_data.pop(sk)
                break;
        else:
            ws3_count += 1
            ws3['A' + str(ws3_count)] = mk
            ws3['B' + str(ws3_count)] = mv
    for sk, sv in sub_data.items():
        for i in sv:
            ws4_count += 1
            ws4['A' + str(ws4_count)] = i
        ws4_count += 1
        ws4['A' + str(ws4_count)] = ""

    wb.save("test2.xlsx")
    print("Sorted Count: " + str(ws1_count))


def extract_main(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[0]

    wb_data = dict()

    for r in ws.rows:
        try:
            wb_data[r[0].value.upper()] = r[1].value
        except AttributeError:
            wb_data[r[0].value] = r[1].value

    return wb_data


def extract_data(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws2 = wb.worksheets[1]

    w2_data = dict()
    count = 0
    for r in ws2.rows:
        count += 1
        try:
            r0 = r[0].value.upper()
            r1 = r[1].value.upper()
        except AttributeError as e:
            print(e)
            continue;

        if r0 in w2_data:
            w2_data[r0].append(r1)
        else:
            w2_data[r0] = [r0, r1]

    for k, v in w2_data.items():
        w2_data[k] = list(set(w2_data[k]))
    print(count)
    return w2_data


def main():
    parser = argparse.ArgumentParser(description="Crawling Data Analyzer")

    # I&O file
    parser.add_argument('--excel_path', dest="excel_path", type=str, default="./data/data2.xlsx",
                        help="Input file path")

    args = parser.parse_args()

    w1_data = extract_main(args.excel_path)
    w2_data = extract_data(args.excel_path)

    print("Main data length: ", len(w1_data))
    print("Sub data length: ", len(w2_data))

    compare_data(w1_data, w2_data)


if __name__ == '__main__':
    main()
